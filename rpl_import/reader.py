# -*- coding: utf-8 -*-
"""Source readers: Excel workbook / CSV -> :class:`SourceGrid`.

A :class:`SourceGrid` is a plain in-memory rectangle of cell values with
1-based ``cell(row, col)`` access. Merged Excel ranges are resolved by
replicating the anchor value across the range so multi-row/merged headers
read naturally. Values come from ``openpyxl(data_only=True)``, i.e. formula
cells yield their last *cached* result (or ``None`` when the workbook was
saved without one) — :func:`SourceGrid.formula_gaps` lets callers surface
that honestly instead of silently importing blanks.

No qgis/Qt imports. openpyxl is imported lazily so this module always
imports; loading a workbook without openpyxl raises :class:`ReaderError`
with an actionable message.
"""

from __future__ import annotations

import csv
import io
import os
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

# Also cover direct use of this pure package without plugin package startup.
os.environ["OPENPYXL_LXML"] = "False"

EXCEL_EXTENSIONS = (".xlsx", ".xlsm")
CSV_EXTENSIONS = (".csv", ".txt")

#: Detection reads at most this many rows per sheet; the full parse is bounded
#: by the confirmed data range instead.
DETECTION_SAMPLE_ROWS = 400

#: Hard ceiling on rows read per sheet (10k positions in an alternating layout
#: is ~20k rows; leave generous headroom while bounding memory).
MAX_ROWS = 200_000
MAX_COLS = 256


class ReaderError(Exception):
    """Source file could not be read; message is user-facing."""


@dataclass
class SourceGrid:
    """One worksheet (or CSV) as a value rectangle with provenance."""
    sheet: str
    rows: List[List[object]] = field(default_factory=list)  # rows[r-1][c-1]
    n_cols: int = 0
    #: (row, col) -> formula text for cells whose cached value was missing.
    uncached_formulas: Dict[Tuple[int, int], str] = field(default_factory=dict)
    truncated: bool = False

    @property
    def n_rows(self) -> int:
        return len(self.rows)

    def cell(self, row: int, col: int):
        """1-based access; out-of-range reads return None."""
        if 1 <= row <= len(self.rows):
            row_values = self.rows[row - 1]
            if 1 <= col <= len(row_values):
                return row_values[col - 1]
        return None

    def row_values(self, row: int) -> List[object]:
        if 1 <= row <= len(self.rows):
            values = list(self.rows[row - 1])
            values.extend([None] * (self.n_cols - len(values)))
            return values
        return [None] * self.n_cols

    def formula_gaps(self, start_row: int, end_row: int) -> List[Tuple[int, int]]:
        """(row, col) of uncached formula cells inside the given row range."""
        return sorted(
            (r, c) for (r, c) in self.uncached_formulas if start_row <= r <= end_row
        )


def _require_openpyxl():
    try:
        import openpyxl
    except Exception as exc:  # pragma: no cover - environment specific
        raise ReaderError(
            "openpyxl is required to read Excel files but could not be "
            "imported. Ensure the plugin's lib/ folder is present and not "
            "blocked by antivirus. (%s)" % exc
        )
    if getattr(openpyxl, "LXML", False):
        raise ReaderError(
            "Excel support was loaded with an unsafe native XML backend. "
            "Restart QGIS once to activate the safe workbook reader, then "
            "try the import again."
        )
    return openpyxl.load_workbook


def is_excel(path: str) -> bool:
    return os.path.splitext(path)[1].lower() in EXCEL_EXTENSIONS


def is_csv(path: str) -> bool:
    return os.path.splitext(path)[1].lower() in CSV_EXTENSIONS


def sheet_names(path: str) -> List[str]:
    """Worksheet names of an Excel file (a CSV counts as one sheet)."""
    if is_csv(path):
        return [os.path.splitext(os.path.basename(path))[0]]
    load_workbook = _require_openpyxl()
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ReaderError(f"Could not open '{os.path.basename(path)}': {exc}")
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def load_grid(path: str, sheet: Optional[str] = None,
              max_rows: int = MAX_ROWS) -> SourceGrid:
    """Load one worksheet (or the CSV) fully, merged ranges resolved."""
    if is_csv(path):
        return _load_csv_grid(path, max_rows=max_rows)
    grids = _load_excel_grids(path, only_sheet=sheet, max_rows=max_rows)
    if not grids:
        raise ReaderError(
            f"Sheet '{sheet}' not found in '{os.path.basename(path)}'."
            if sheet else f"No worksheets found in '{os.path.basename(path)}'."
        )
    return grids[0]


def load_sample_grids(path: str,
                      max_rows: int = DETECTION_SAMPLE_ROWS) -> List[SourceGrid]:
    """Bounded-row grids for every sheet — worksheet scanning/scoring."""
    if is_csv(path):
        return [_load_csv_grid(path, max_rows=max_rows)]
    return _load_excel_grids(path, only_sheet=None, max_rows=max_rows)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def _load_excel_grids(path: str, only_sheet: Optional[str],
                      max_rows: int) -> List[SourceGrid]:
    load_workbook = _require_openpyxl()
    try:
        # Full (non read-only) load: needed for merged-cell ranges. data_only
        # gives cached formula results; the second lightweight pass below
        # records which formula cells had no cached value.
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:
        raise ReaderError(f"Could not open '{os.path.basename(path)}': {exc}")
    try:
        formula_cells = _formula_cells(load_workbook, path, only_sheet, max_rows)
        grids: List[SourceGrid] = []
        for name in workbook.sheetnames:
            if only_sheet is not None and name != only_sheet:
                continue
            ws = workbook[name]
            n_rows = min(ws.max_row or 0, max_rows)
            n_cols = min(ws.max_column or 0, MAX_COLS)
            rows: List[List[object]] = []
            for row in ws.iter_rows(min_row=1, max_row=n_rows,
                                    min_col=1, max_col=n_cols,
                                    values_only=True):
                rows.append(list(row))
            grid = SourceGrid(
                sheet=name, rows=rows, n_cols=n_cols,
                truncated=bool((ws.max_row or 0) > max_rows),
            )
            _apply_merged_ranges(grid, ws, n_rows, n_cols)
            for (r, c), formula in (formula_cells.get(name) or {}).items():
                if r <= n_rows and c <= n_cols and grid.cell(r, c) is None:
                    grid.uncached_formulas[(r, c)] = formula
            grids.append(grid)
        return grids
    finally:
        workbook.close()


def _apply_merged_ranges(grid: SourceGrid, ws, n_rows: int, n_cols: int) -> None:
    """Replicate each merged range's anchor value across the range."""
    try:
        merged = list(ws.merged_cells.ranges)
    except Exception:
        return
    for rng in merged:
        anchor = grid.cell(rng.min_row, rng.min_col)
        if anchor is None:
            continue
        for r in range(rng.min_row, min(rng.max_row, n_rows) + 1):
            for c in range(rng.min_col, min(rng.max_col, n_cols) + 1):
                if 1 <= r <= len(grid.rows) and 1 <= c <= len(grid.rows[r - 1]):
                    if grid.rows[r - 1][c - 1] is None:
                        grid.rows[r - 1][c - 1] = anchor


def _formula_cells(load_workbook, path: str, only_sheet: Optional[str],
                   max_rows: int) -> Dict[str, Dict[Tuple[int, int], str]]:
    """Formula text per sheet, read via a cheap read-only pass.

    Returns {} when the second pass fails — formula provenance is best-effort
    and must never block an import on its own.
    """
    result: Dict[str, Dict[Tuple[int, int], str]] = {}
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception:
        return result
    try:
        for name in workbook.sheetnames:
            if only_sheet is not None and name != only_sheet:
                continue
            ws = workbook[name]
            cells: Dict[Tuple[int, int], str] = {}
            for r, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, max_rows),
                                 min_col=1, max_col=min(ws.max_column or 0, MAX_COLS),
                                 values_only=True), start=1):
                for c, value in enumerate(row, start=1):
                    if isinstance(value, str) and value.startswith("="):
                        cells[(r, c)] = value
            if cells:
                result[name] = cells
    except Exception:
        return {}
    finally:
        try:
            workbook.close()
        except Exception:
            pass
    return result


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
def _load_csv_grid(path: str, max_rows: int) -> SourceGrid:
    try:
        with io.open(path, "r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(64 * 1024)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel
            reader = csv.reader(handle, dialect)
            rows: List[List[object]] = []
            truncated = False
            for i, row in enumerate(reader):
                if i >= max_rows:
                    truncated = True
                    break
                rows.append([_csv_value(cell) for cell in row])
    except OSError as exc:
        raise ReaderError(f"Could not open '{os.path.basename(path)}': {exc}")
    n_cols = min(max((len(r) for r in rows), default=0), MAX_COLS)
    rows = [r[:n_cols] for r in rows]
    return SourceGrid(
        sheet=os.path.splitext(os.path.basename(path))[0],
        rows=rows, n_cols=n_cols, truncated=truncated,
    )


def _csv_value(text: str) -> object:
    """CSV cells arrive as text; give numbers back their type."""
    stripped = (text or "").strip()
    if stripped == "":
        return None
    try:
        return int(stripped)
    except ValueError:
        pass
    try:
        return float(stripped)
    except ValueError:
        return stripped


# ---------------------------------------------------------------------------
# File fingerprint (audit)
# ---------------------------------------------------------------------------
def file_fingerprint(path: str) -> Dict[str, object]:
    """SHA-256 + size + mtime for the import audit record."""
    import hashlib

    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    stat = os.stat(path)
    return {
        "path": os.path.abspath(path),
        "filename": os.path.basename(path),
        "sha256": digest.hexdigest(),
        "size_bytes": stat.st_size,
        "modified_utc": __import__("time").strftime(
            "%Y-%m-%dT%H:%M:%SZ", __import__("time").gmtime(stat.st_mtime)),
    }
