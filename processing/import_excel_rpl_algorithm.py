# import_excel_rpl.py
# -*- coding: utf-8 -*-
"""
ImportExcelRPL
Import Excel Cable RPLs to QGIS

This version accepts column letters (e.g. "A", "B", "C", etc.) as input.

Points vs Lines:
  - Points fields: PosNo, Event, DistCumulative, CableDistCumulative, ApproxDepth, Remarks,
                   Latitude, Longitude, SourceFile
  - Lines fields : FromPos, ToPos, Bearing, DistBetweenPos, Slack, CableDistBetweenPos, CableCode,
                   FiberPair, CableType, LayDirection, LayVessel, ProtectionMethod, DateInstalled,
                   TargetBurialDepth, BurialDepth, TerritorialWater, EEZ, SourceFile
"""

__author__ = 'Kieran McMonagle'
__date__ = '2024-08-29'
__copyright__ = '(C) 2024 by Kieran McMonagle'
__revision__ = '$Format:%H$'

import os
import sys
plugin_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
lib_dir = os.path.join(plugin_dir, 'lib')
if os.path.isdir(lib_dir) and lib_dir not in sys.path:
    sys.path.insert(0, lib_dir)

from qgis.PyQt.QtCore import QCoreApplication, QVariant, QSettings
from qgis.core import (
    QgsProcessing,
    QgsProcessingAlgorithm,
    QgsProcessingParameterFile,
    QgsProcessingParameterString,
    QgsProcessingParameterNumber,
    QgsProcessingParameterFeatureSink,
    QgsProcessingException,
    QgsFeatureSink,
    QgsFeature,
    QgsFields,
    QgsField,
    QgsPointXY,
    QgsGeometry,
    QgsWkbTypes,
    QgsCoordinateReferenceSystem,
    QgsProcessingLayerPostProcessorInterface
)

try:
    # openpyxl is bundled under plugin lib/ but may also be available in QGIS python
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

def col_letter_to_index(letter):
    """
    Convert Excel column letter(s) to a 1-based column index.
    For example, 'A' -> 1, 'B' -> 2, ... 'Z' -> 26, 'AA' -> 27, etc.
    Returns 0 for empty or "0".
    """
    if not letter or letter.strip() == "0":
        return 0
    letter = letter.strip().upper()
    result = 0
    for char in letter:
        if not ('A' <= char <= 'Z'):
            return 0
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result

class ImportExcelRPLAlgorithm(QgsProcessingAlgorithm):
    """
    A QGIS Processing Algorithm that:
      - Prompts the user for an Excel file, sheet name, start row and optionally an end row.
      - Prompts for column letters for the mandatory and additional fields.
      - Interprets alternating rows as Point and Line features.
      - Outputs two memory layers: one for Points and one for Lines.
    """

    # --- Parameter identifiers ---
    INPUT_EXCEL    = 'INPUT_EXCEL'
    INPUT_SHEET    = 'INPUT_SHEET'
    INPUT_STARTROW = 'INPUT_STARTROW'
    INPUT_ENDROW   = 'INPUT_ENDROW'
    INFO           = 'INFO'

    # Mandatory columns for Points geometry (as letters)
    COL_POSNO   = 'COL_POSNO'
    COL_EVENT   = 'COL_EVENT'
    COL_LATDEG  = 'COL_LATDEG'
    COL_LATMIN  = 'COL_LATMIN'
    COL_LATHEMI = 'COL_LATHEMI'
    COL_LONDEG  = 'COL_LONDEG'
    COL_LONMIN  = 'COL_LONMIN'
    COL_LONHEMI = 'COL_LONHEMI'

    # Additional fields
    COL_BEARING           = 'COL_BEARING'
    COL_DISTBETWEENPOS    = 'COL_DISTBETWEENPOS'
    COL_DISTCUMUL         = 'COL_DISTCUMUL'
    COL_SLACK             = 'COL_SLACK'
    COL_CABLEDISTBETWEEN  = 'COL_CABLEDISTBETWEEN'
    COL_CABLEDISTCUM      = 'COL_CABLEDISTCUM'
    COL_CABLECODE         = 'COL_CABLECODE'
    COL_FIBERPAIR         = 'COL_FIBERPAIR'
    COL_CABLETYPE         = 'COL_CABLETYPE'
    COL_APPROXDEPTH       = 'COL_APPROXDEPTH'
    COL_LAYDIRECTION      = 'COL_LAYDIRECTION'
    COL_LAYVESSEL         = 'COL_LAYVESSEL'
    COL_PROTECTIONMETHOD  = 'COL_PROTECTIONMETHOD'
    COL_DATEINSTALLED     = 'COL_DATEINSTALLED'
    COL_TARGETBURIALDEPTH = 'COL_TARGETBURIALDEPTH'
    COL_BURIALDEPTH       = 'COL_BURIALDEPTH'
    COL_REMARKS           = 'COL_REMARKS'
    COL_CHARTNO           = 'COL_CHARTNO'
    COL_TERRWATER         = 'COL_TERRWATER'
    COL_EEZ               = 'COL_EEZ'

    # Outputs
    OUTPUT_POINTS = 'OUTPUT_POINTS'
    OUTPUT_LINES  = 'OUTPUT_LINES'

    def tr(self, string):
        return QCoreApplication.translate('Processing', string)

    def createInstance(self):
        return ImportExcelRPLAlgorithm()

    def name(self):
        return 'importexcelrpl'

    def displayName(self):
        return self.tr('Import Excel RPL')

    def group(self):
        # Place this algorithm in the "RPL Tools" group
        return self.tr('RPL Tools')

    def groupId(self):
        return 'rpl_tools'

    def shortHelpString(self):
        return self.tr("""
<h3>Import Excel RPL</h3>
<p>This tool imports a cable Route Position List (RPL) from a Microsoft Excel file (.xlsx) into QGIS. It creates two new layers: a <b>point layer</b> for positions (e.g., Start of Cable, End of Cable, Repeater) and a <b>line layer</b> representing the cable segments connecting these points.</p>

<h4>How it Works</h4>
<p>The tool reads the specified sheet in the Excel file row by row, interpreting alternating rows as point and line data.
<ul>
  <li>The first data row (specified by 'Data Start Row') is treated as a <b>Point</b>.</li>
  <li>The second row is treated as a <b>Line</b> (connecting the first and second points).</li>
  <li>The third row is another <b>Point</b>.</li>
  <li>The fourth row is another <b>Line</b> (connecting the second and third points), and so on.</li>
</ul>
This pattern continues until the 'Data End Row' is reached or the end of the sheet.
</p>

<h4>Input Parameters</h4>
<ul>
  <li><b>Select Excel RPL File:</b> The Excel file (.xlsx) containing the RPL.</li>
  <li><b>Sheet Name:</b> The exact name of the worksheet with the RPL data.</li>
  <li><b>Data Start Row:</b> The row number where the RPL data begins.</li>
  <li><b>Data End Row:</b> (Optional) The last row of data to import. If set to 0, it reads to the end of the sheet.</li>
  <li><b>Column IDs:</b> You must provide the Excel column letters (e.g., A, B, AA) for each required data field.
    <ul>
      <li><b>Mandatory for Geometry:</b> The tool requires columns for Latitude and Longitude, split into Degrees, Minutes, and Hemisphere (e.g., LatDeg, LatMin, LatHemi). Without these, points cannot be created.</li>
      <li><b>Other Fields:</b> All other fields are optional. If a column letter is not provided for an optional field (or set to '0'), it will be skipped.</li>
    </ul>
  </li>
</ul>

<h4>Outputs</h4>
<ol>
  <li><b>RPL Points:</b> A point layer with features for each point row in the RPL. The layer's CRS will be WGS 84 (EPSG:4326).</li>
  <li><b>RPL Lines:</b> A line layer with features for each line segment connecting two consecutive points. The layer's CRS will also be WGS 84 (EPSG:4326).</li>
</ol>

<h4>Notes</h4>
<ul>
  <li>The tool expects latitude and longitude in Degrees, Decimal Minutes, and Hemisphere format. It converts these to Decimal Degrees automatically.</li>
  <li>If a row that should define a point has invalid coordinate data, it will be skipped, and a warning will be shown in the Log Messages Panel. This may affect the creation of subsequent line features.</li>
  <li>The 'SourceFile' attribute is automatically added to both output layers to maintain traceability to the original Excel file.</li>
</ul>
""")

    def initAlgorithm(self, config=None):
        # --- Read settings ---
        settings = QSettings()
        
        # Define default column mappings
        default_mappings = {
            self.COL_POSNO: "A", self.COL_EVENT: "B", self.COL_LATDEG: "C",
            self.COL_LATMIN: "D", self.COL_LATHEMI: "E", self.COL_LONDEG: "F",
            self.COL_LONMIN: "G", self.COL_LONHEMI: "H", self.COL_BEARING: "I",
            self.COL_DISTBETWEENPOS: "J", self.COL_DISTCUMUL: "K", self.COL_SLACK: "L",
            self.COL_CABLEDISTBETWEEN: "M", self.COL_CABLEDISTCUM: "N",
            self.COL_CABLECODE: "O", self.COL_FIBERPAIR: "P", self.COL_CABLETYPE: "Q",
            self.COL_APPROXDEPTH: "T", self.COL_LAYDIRECTION: "U",
            self.COL_LAYVESSEL: "V", self.COL_PROTECTIONMETHOD: "W",
            self.COL_DATEINSTALLED: "X", self.COL_TARGETBURIALDEPTH: "Y",
            self.COL_BURIALDEPTH: "0", self.COL_REMARKS: "Z", self.COL_CHARTNO: "0",
            self.COL_TERRWATER: "AA", self.COL_EEZ: "AB"
        }

        # Load saved mappings or use defaults
        saved_mappings = {}
        for key, default_val in default_mappings.items():
            saved_mappings[key] = settings.value(f"SubseaCableTools/ImportExcelRPL/{key}", default_val)

        # 1) Excel File
        self.addParameter(
            QgsProcessingParameterFile(
                self.INPUT_EXCEL,
                self.tr('Select Excel RPL File'),
                extension='xlsx'
            )
        )
        # 2) Sheet Name
        self.addParameter(
            QgsProcessingParameterString(
                self.INPUT_SHEET,
                self.tr('Sheet Name (e.g. RPLSheet)')
            )
        )
        # 3) Start Row
        self.addParameter(
            QgsProcessingParameterNumber(
                self.INPUT_STARTROW,
                self.tr('Data Start Row'),
                type=QgsProcessingParameterNumber.Integer,
                defaultValue=7
            )
        )
        # 4) End Row (optional; 0 = no limit)
        self.addParameter(
            QgsProcessingParameterNumber(
                self.INPUT_ENDROW,
                self.tr('Data End Row (0 for no limit)'),
                type=QgsProcessingParameterNumber.Integer,
                defaultValue=0,
                optional=True
            )
        )
        # 5) Instruction (info text)
        info_param = QgsProcessingParameterString(
            self.INFO,
            self.tr('Input the column IDs below (enter a letter, e.g. A, B, ...)'),
            defaultValue="",
            optional=True
        )
        info_param.setFlags(info_param.flags() | QgsProcessingParameterString.FlagAdvanced)
        self.addParameter(info_param)

        # -- Mandatory columns for Points --
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_POSNO,
                self.tr('PosNo'),
                defaultValue=saved_mappings[self.COL_POSNO]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_EVENT,
                self.tr('Event'),
                defaultValue=saved_mappings[self.COL_EVENT]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_LATDEG,
                self.tr('LatDeg'),
                defaultValue=saved_mappings[self.COL_LATDEG]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_LATMIN,
                self.tr('LatMin'),
                defaultValue=saved_mappings[self.COL_LATMIN]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_LATHEMI,
                self.tr('LatHemi'),
                defaultValue=saved_mappings[self.COL_LATHEMI]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_LONDEG,
                self.tr('LonDeg'),
                defaultValue=saved_mappings[self.COL_LONDEG]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_LONMIN,
                self.tr('LonMin'),
                defaultValue=saved_mappings[self.COL_LONMIN]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_LONHEMI,
                self.tr('LonHemi'),
                defaultValue=saved_mappings[self.COL_LONHEMI]
            )
        )

        # -- Additional fields --
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_BEARING,
                self.tr('Bearing'),
                defaultValue=saved_mappings[self.COL_BEARING]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_DISTBETWEENPOS,
                self.tr('DistBetweenPos'),
                defaultValue=saved_mappings[self.COL_DISTBETWEENPOS]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_DISTCUMUL,
                self.tr('DistCumulative'),
                defaultValue=saved_mappings[self.COL_DISTCUMUL]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_SLACK,
                self.tr('Slack'),
                defaultValue=saved_mappings[self.COL_SLACK]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_CABLEDISTBETWEEN,
                self.tr('CableDistBetweenPos'),
                defaultValue=saved_mappings[self.COL_CABLEDISTBETWEEN]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_CABLEDISTCUM,
                self.tr('CableDistCumulative'),
                defaultValue=saved_mappings[self.COL_CABLEDISTCUM]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_CABLECODE,
                self.tr('CableCode'),
                defaultValue=saved_mappings[self.COL_CABLECODE]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_FIBERPAIR,
                self.tr('FiberPair'),
                defaultValue=saved_mappings[self.COL_FIBERPAIR]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_CABLETYPE,
                self.tr('CableType'),
                defaultValue=saved_mappings[self.COL_CABLETYPE]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_APPROXDEPTH,
                self.tr('ApproxDepth'),
                defaultValue=saved_mappings[self.COL_APPROXDEPTH]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_LAYDIRECTION,
                self.tr('LayDirection'),
                defaultValue=saved_mappings[self.COL_LAYDIRECTION]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_LAYVESSEL,
                self.tr('LayVessel'),
                defaultValue=saved_mappings[self.COL_LAYVESSEL]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_PROTECTIONMETHOD,
                self.tr('ProtectionMethod'),
                defaultValue=saved_mappings[self.COL_PROTECTIONMETHOD]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_DATEINSTALLED,
                self.tr('DateInstalled'),
                defaultValue=saved_mappings[self.COL_DATEINSTALLED]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_TARGETBURIALDEPTH,
                self.tr('TargetBurialDepth'),
                defaultValue=saved_mappings[self.COL_TARGETBURIALDEPTH]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_BURIALDEPTH,
                self.tr('BurialDepth'),
                defaultValue=saved_mappings[self.COL_BURIALDEPTH]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_REMARKS,
                self.tr('Remarks'),
                defaultValue=saved_mappings[self.COL_REMARKS]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_CHARTNO,
                self.tr('Chart No'),
                defaultValue=saved_mappings[self.COL_CHARTNO]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_TERRWATER,
                self.tr('TerritorialWater'),
                defaultValue=saved_mappings[self.COL_TERRWATER]
            )
        )
        self.addParameter(
            QgsProcessingParameterString(
                self.COL_EEZ,
                self.tr('EEZ'),
                defaultValue=saved_mappings[self.COL_EEZ]
            )
        )

        # --- Outputs ---
        self.addParameter(
            QgsProcessingParameterFeatureSink(
                self.OUTPUT_POINTS,
                self.tr('RPL Points')
            )
        )
        self.addParameter(
            QgsProcessingParameterFeatureSink(
                self.OUTPUT_LINES,
                self.tr('RPL Lines')
            )
        )

    def processAlgorithm(self, parameters, context, feedback):
        # Retrieve inputs
        excel_file = self.parameterAsFile(parameters, self.INPUT_EXCEL, context)
        sheet_name = self.parameterAsString(parameters, self.INPUT_SHEET, context)
        start_row  = self.parameterAsInt(parameters, self.INPUT_STARTROW, context)
        end_row    = self.parameterAsInt(parameters, self.INPUT_ENDROW, context)

        # Store filename (without extension) for dynamic layer naming
        filename = os.path.splitext(os.path.basename(excel_file))[0]

        # --- Save settings ---
        settings = QSettings()
        param_keys = [
            self.COL_POSNO, self.COL_EVENT, self.COL_LATDEG, self.COL_LATMIN,
            self.COL_LATHEMI, self.COL_LONDEG, self.COL_LONMIN, self.COL_LONHEMI,
            self.COL_BEARING, self.COL_DISTBETWEENPOS, self.COL_DISTCUMUL,
            self.COL_SLACK, self.COL_CABLEDISTBETWEEN, self.COL_CABLEDISTCUM,
            self.COL_CABLECODE, self.COL_FIBERPAIR, self.COL_CABLETYPE,
            self.COL_APPROXDEPTH, self.COL_LAYDIRECTION, self.COL_LAYVESSEL,
            self.COL_PROTECTIONMETHOD, self.COL_DATEINSTALLED,
            self.COL_TARGETBURIALDEPTH, self.COL_BURIALDEPTH, self.COL_REMARKS,
            self.COL_CHARTNO, self.COL_TERRWATER, self.COL_EEZ
        ]
        for key in param_keys:
            value = self.parameterAsString(parameters, key, context)
            settings.setValue(f"SubseaCableTools/ImportExcelRPL/{key}", value)

        # Extract file name for traceability
        source_file_name = os.path.basename(excel_file)

        # Convert column letters to numeric indexes
        p_col_posno   = col_letter_to_index(self.parameterAsString(parameters, self.COL_POSNO, context))
        p_col_event   = col_letter_to_index(self.parameterAsString(parameters, self.COL_EVENT, context))
        p_col_latdeg  = col_letter_to_index(self.parameterAsString(parameters, self.COL_LATDEG, context))
        p_col_latmin  = col_letter_to_index(self.parameterAsString(parameters, self.COL_LATMIN, context))
        p_col_lathemi = col_letter_to_index(self.parameterAsString(parameters, self.COL_LATHEMI, context))
        p_col_londeg  = col_letter_to_index(self.parameterAsString(parameters, self.COL_LONDEG, context))
        p_col_lonmin  = col_letter_to_index(self.parameterAsString(parameters, self.COL_LONMIN, context))
        p_col_lonhemi = col_letter_to_index(self.parameterAsString(parameters, self.COL_LONHEMI, context))

        p_col_bearing          = col_letter_to_index(self.parameterAsString(parameters, self.COL_BEARING, context))
        p_col_distbetweenpos   = col_letter_to_index(self.parameterAsString(parameters, self.COL_DISTBETWEENPOS, context))
        p_col_distcumul        = col_letter_to_index(self.parameterAsString(parameters, self.COL_DISTCUMUL, context))
        p_col_slack            = col_letter_to_index(self.parameterAsString(parameters, self.COL_SLACK, context))
        p_col_cabledistbetween = col_letter_to_index(self.parameterAsString(parameters, self.COL_CABLEDISTBETWEEN, context))
        p_col_cabledistcum     = col_letter_to_index(self.parameterAsString(parameters, self.COL_CABLEDISTCUM, context))
        p_col_cablecode        = col_letter_to_index(self.parameterAsString(parameters, self.COL_CABLECODE, context))
        p_col_fiberpair        = col_letter_to_index(self.parameterAsString(parameters, self.COL_FIBERPAIR, context))
        p_col_cabletype        = col_letter_to_index(self.parameterAsString(parameters, self.COL_CABLETYPE, context))
        p_col_approxdepth      = col_letter_to_index(self.parameterAsString(parameters, self.COL_APPROXDEPTH, context))
        p_col_laydirection     = col_letter_to_index(self.parameterAsString(parameters, self.COL_LAYDIRECTION, context))
        p_col_layvessel        = col_letter_to_index(self.parameterAsString(parameters, self.COL_LAYVESSEL, context))
        p_col_protectionmethod = col_letter_to_index(self.parameterAsString(parameters, self.COL_PROTECTIONMETHOD, context))
        p_col_dateinstalled    = col_letter_to_index(self.parameterAsString(parameters, self.COL_DATEINSTALLED, context))
        p_col_targetburialdepth= col_letter_to_index(self.parameterAsString(parameters, self.COL_TARGETBURIALDEPTH, context))
        p_col_burialdepth      = col_letter_to_index(self.parameterAsString(parameters, self.COL_BURIALDEPTH, context))
        p_col_remarks          = col_letter_to_index(self.parameterAsString(parameters, self.COL_REMARKS, context))
        p_col_chartno          = col_letter_to_index(self.parameterAsString(parameters, self.COL_CHARTNO, context))
        p_col_terrwater        = col_letter_to_index(self.parameterAsString(parameters, self.COL_TERRWATER, context))
        p_col_eez              = col_letter_to_index(self.parameterAsString(parameters, self.COL_EEZ, context))

        # Prepare storage for point and line rows
        # Open workbook with robust error handling
        if load_workbook is None:
            raise QgsProcessingException(
                'openpyxl is required to read Excel files but could not be imported. '
                "If you're running this from the plugin, ensure the plugin's lib/ folder is present and not blocked by antivirus."
            )
        try:
            wb = load_workbook(excel_file, data_only=True)
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            raise QgsProcessingException(
                f"Failed to open Excel file '{excel_file}'. This may be due to file corruption, unsupported format, or a library issue.\nError: {e}\nTraceback:\n{tb}"
            )
        if sheet_name not in wb.sheetnames:
            raise QgsProcessingException(
                f"Sheet '{sheet_name}' not found in {excel_file}. Available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]

        # Determine effective last row
        workbook_max_row = ws.max_row
        if end_row and end_row > 0:
            effective_max_row = min(workbook_max_row, end_row)
        else:
            effective_max_row = workbook_max_row

        # Define fields for Points
        point_fields = QgsFields()
        point_fields.append(QgsField("PosNo", QVariant.Int))
        point_fields.append(QgsField("Event", QVariant.String))
        point_fields.append(QgsField("DistCumulative", QVariant.Double))
        point_fields.append(QgsField("CableDistCumulative", QVariant.Double))
        point_fields.append(QgsField("ApproxDepth", QVariant.Double))
        point_fields.append(QgsField("Remarks", QVariant.String))
        point_fields.append(QgsField("ChartNo", QVariant.Int))
        point_fields.append(QgsField("Latitude", QVariant.Double))
        point_fields.append(QgsField("Longitude", QVariant.Double))
        point_fields.append(QgsField("SourceFile", QVariant.String))

        # Define fields for Lines
        line_fields = QgsFields()
        line_fields.append(QgsField("FromPos", QVariant.Int))
        line_fields.append(QgsField("ToPos", QVariant.Int))
        line_fields.append(QgsField("Bearing", QVariant.Double))
        line_fields.append(QgsField("DistBetweenPos", QVariant.Double))
        line_fields.append(QgsField("Slack", QVariant.Double))
        line_fields.append(QgsField("CableDistBetweenPos", QVariant.Double))
        line_fields.append(QgsField("CableCode", QVariant.String))
        line_fields.append(QgsField("FiberPair", QVariant.String))
        line_fields.append(QgsField("CableType", QVariant.String))
        line_fields.append(QgsField("LayDirection", QVariant.String))
        line_fields.append(QgsField("LayVessel", QVariant.String))
        line_fields.append(QgsField("ProtectionMethod", QVariant.String))
        line_fields.append(QgsField("DateInstalled", QVariant.String))
        line_fields.append(QgsField("TargetBurialDepth", QVariant.Double))
        line_fields.append(QgsField("BurialDepth", QVariant.Double))
        line_fields.append(QgsField("TerritorialWater", QVariant.String))
        line_fields.append(QgsField("EEZ", QVariant.String))
        line_fields.append(QgsField("SourceFile", QVariant.String))

        # --- Sinks for output layers ---
        # Get sinks for the two output layers
        (point_sink, point_sink_id) = self.parameterAsSink(
            parameters,
            self.OUTPUT_POINTS,
            context,
            point_fields,
            QgsWkbTypes.Point,
            QgsCoordinateReferenceSystem("EPSG:4326")
        )
        (line_sink, line_sink_id) = self.parameterAsSink(
            parameters,
            self.OUTPUT_LINES,
            context,
            line_fields,
            QgsWkbTypes.LineString,
            QgsCoordinateReferenceSystem("EPSG:4326")
        )

        # Helper functions for latitude/longitude parsing
        def parse_lat(row_idx):
            if (p_col_latdeg <= 0) or (p_col_latmin <= 0) or (p_col_lathemi <= 0):
                return None
            deg_val  = ws.cell(row=row_idx, column=p_col_latdeg).value
            min_val  = ws.cell(row=row_idx, column=p_col_latmin).value
            hemi_val = ws.cell(row=row_idx, column=p_col_lathemi).value
            if deg_val is None or min_val is None or not hemi_val:
                return None
            try:
                deg = abs(float(deg_val))
                mins = abs(float(min_val))
            except:
                return None
            lat = deg + (mins / 60.0)
            if str(hemi_val).strip().upper().startswith('S'):
                lat = -lat
            return lat

        def parse_lon(row_idx):
            if (p_col_londeg <= 0) or (p_col_lonmin <= 0) or (p_col_lonhemi <= 0):
                return None
            deg_val  = ws.cell(row=row_idx, column=p_col_londeg).value
            min_val  = ws.cell(row=row_idx, column=p_col_lonmin).value
            hemi_val = ws.cell(row=row_idx, column=p_col_lonhemi).value
            if deg_val is None or min_val is None or not hemi_val:
                return None
            try:
                deg = abs(float(deg_val))
                mins = abs(float(min_val))
            except:
                return None
            lon = deg + (mins / 60.0)
            if str(hemi_val).strip().upper().startswith('W'):
                lon = -lon
            return lon
        
        # Initialize processing variables
        processed = 0
        row_count = effective_max_row - start_row + 1
        point_list = []
        line_list = []
        
        # Alternate: even-indexed rows = Points, odd-indexed = Lines.
        consecutive_invalid_rows = 0
        max_consecutive_invalid = 3  # Stop after 3 consecutive invalid rows
        
        for row_idx in range(start_row, effective_max_row + 1):
            if feedback.isCanceled():
                break
            processed += 1
            feedback.setProgress(int(100 * processed / row_count))
            offset = row_idx - start_row
            
            # Check if this appears to be end of RPL data
            if offset % 2 == 0:
                # Point row - check if coordinates are valid
                lat_dd = parse_lat(row_idx)
                lon_dd = parse_lon(row_idx)
                if lat_dd is None or lon_dd is None:
                    consecutive_invalid_rows += 1
                    feedback.pushWarning(f"Row {row_idx}: Invalid lat/lon coordinates")
                    
                    # If we've hit multiple consecutive invalid coordinate rows, 
                    # this likely indicates end of RPL data
                    if consecutive_invalid_rows >= max_consecutive_invalid:
                        feedback.pushInfo(f"Detected end of RPL data at row {row_idx} (consecutive invalid coordinates). Processing stopped.")
                        break
                    continue
                else:
                    # Valid coordinates found, reset counter
                    consecutive_invalid_rows = 0
                pos_no = -1
                if p_col_posno > 0:
                    val = ws.cell(row=row_idx, column=p_col_posno).value
                    try:
                        pos_no = int(val) if val is not None else -1
                    except (ValueError, TypeError):
                        pos_no = -1
                        
                event_str = ""
                if p_col_event > 0:
                    val = ws.cell(row=row_idx, column=p_col_event).value
                    event_str = str(val) if val else ""
                    
                dist_cumul = None
                if p_col_distcumul > 0:
                    val = ws.cell(row=row_idx, column=p_col_distcumul).value
                    try:
                        dist_cumul = float(val) if val is not None else None
                    except (ValueError, TypeError):
                        dist_cumul = None
                        
                cable_dist_cumul = None
                if p_col_cabledistcum > 0:
                    val = ws.cell(row=row_idx, column=p_col_cabledistcum).value
                    try:
                        cable_dist_cumul = float(val) if val is not None else None
                    except (ValueError, TypeError):
                        cable_dist_cumul = None
                        
                approx_depth = None
                if p_col_approxdepth > 0:
                    val = ws.cell(row=row_idx, column=p_col_approxdepth).value
                    try:
                        approx_depth = float(val) if val is not None else None
                    except (ValueError, TypeError):
                        approx_depth = None
                        
                remarks_str = ""
                if p_col_remarks > 0:
                    val = ws.cell(row=row_idx, column=p_col_remarks).value
                    remarks_str = str(val) if val else ""
                
                chart_no = 0  # Default value
                if p_col_chartno > 0:
                    val = ws.cell(row=row_idx, column=p_col_chartno).value
                    try:
                        chart_no = int(val) if val is not None else 0
                    except (ValueError, TypeError):
                        chart_no = 0 # Keep default if conversion fails

                point_list.append({
                    "PosNo": pos_no,
                    "Event": event_str,
                    "Lat": lat_dd,
                    "Lon": lon_dd,
                    "DistCumulative": dist_cumul,
                    "CableDistCumulative": cable_dist_cumul,
                    "ApproxDepth": approx_depth,
                    "Remarks": remarks_str,
                    "ChartNo": chart_no
                })
            else:
                # Line row - wrap in try-catch to handle conversion errors
                try:
                    line_attrs = {}
                    conversion_failed = False
                    
                    if p_col_bearing > 0:
                        val = ws.cell(row=row_idx, column=p_col_bearing).value
                        if val is not None:
                            try:
                                line_attrs["Bearing"] = float(val)
                            except (ValueError, TypeError):
                                # If we can't convert bearing to float, this might indicate end of RPL data
                                feedback.pushWarning(f"Row {row_idx}: Could not convert bearing '{val}' to number - possible end of RPL data")
                                conversion_failed = True
                        else:
                            line_attrs["Bearing"] = None
                    
                    if conversion_failed:
                        consecutive_invalid_rows += 1
                        if consecutive_invalid_rows >= max_consecutive_invalid:
                            feedback.pushInfo(f"Detected end of RPL data at row {row_idx} (conversion errors). Processing stopped.")
                            break
                        continue
                    else:
                        consecutive_invalid_rows = 0
                    
                    if p_col_distbetweenpos > 0:
                        val = ws.cell(row=row_idx, column=p_col_distbetweenpos).value
                        try:
                            line_attrs["DistBetweenPos"] = float(val) if val is not None else None
                        except (ValueError, TypeError):
                            line_attrs["DistBetweenPos"] = None
                    
                    if p_col_slack > 0:
                        val = ws.cell(row=row_idx, column=p_col_slack).value
                        try:
                            line_attrs["Slack"] = float(val) if val is not None else None
                        except (ValueError, TypeError):
                            line_attrs["Slack"] = None
                    
                    if p_col_cabledistbetween > 0:
                        val = ws.cell(row=row_idx, column=p_col_cabledistbetween).value
                        try:
                            line_attrs["CableDistBetweenPos"] = float(val) if val is not None else None
                        except (ValueError, TypeError):
                            line_attrs["CableDistBetweenPos"] = None
                    
                    if p_col_cablecode > 0:
                        val = ws.cell(row=row_idx, column=p_col_cablecode).value
                        line_attrs["CableCode"] = str(val) if val else ""
                    if p_col_fiberpair > 0:
                        val = ws.cell(row=row_idx, column=p_col_fiberpair).value
                        line_attrs["FiberPair"] = str(val) if val else ""
                    if p_col_cabletype > 0:
                        val = ws.cell(row=row_idx, column=p_col_cabletype).value
                        line_attrs["CableType"] = str(val) if val else ""
                    if p_col_laydirection > 0:
                        val = ws.cell(row=row_idx, column=p_col_laydirection).value
                        line_attrs["LayDirection"] = str(val) if val else ""
                    if p_col_layvessel > 0:
                        val = ws.cell(row=row_idx, column=p_col_layvessel).value
                        line_attrs["LayVessel"] = str(val) if val else ""
                    if p_col_protectionmethod > 0:
                        val = ws.cell(row=row_idx, column=p_col_protectionmethod).value
                        line_attrs["ProtectionMethod"] = str(val) if val else ""
                    if p_col_dateinstalled > 0:
                        val = ws.cell(row=row_idx, column=p_col_dateinstalled).value
                        line_attrs["DateInstalled"] = str(val) if val else ""
                    if p_col_targetburialdepth > 0:
                        val = ws.cell(row=row_idx, column=p_col_targetburialdepth).value
                        try:
                            line_attrs["TargetBurialDepth"] = float(val) if val is not None else None
                        except (ValueError, TypeError):
                            line_attrs["TargetBurialDepth"] = None
                    if p_col_burialdepth > 0:
                        val = ws.cell(row=row_idx, column=p_col_burialdepth).value
                        try:
                            line_attrs["BurialDepth"] = float(val) if val is not None else None
                        except (ValueError, TypeError):
                            line_attrs["BurialDepth"] = None
                    if p_col_terrwater > 0:
                        val = ws.cell(row=row_idx, column=p_col_terrwater).value
                        line_attrs["TerritorialWater"] = str(val) if val else ""
                    if p_col_eez > 0:
                        val = ws.cell(row=row_idx, column=p_col_eez).value
                        line_attrs["EEZ"] = str(val) if val else ""
                    
                    line_list.append(line_attrs)
                    
                except Exception as e:
                    # If any unexpected error occurs during line processing, 
                    # this might indicate end of RPL data
                    feedback.pushWarning(f"Row {row_idx}: Error processing line data - {str(e)}")
                    consecutive_invalid_rows += 1
                    if consecutive_invalid_rows >= max_consecutive_invalid:
                        feedback.pushInfo(f"Detected end of RPL data at row {row_idx} (processing errors). Processing stopped.")
                        break

        # Create point features
        for pt in point_list:
            feat = QgsFeature(point_fields)
            geom = QgsGeometry.fromPointXY(QgsPointXY(pt["Lon"], pt["Lat"]))
            feat.setGeometry(geom)
            feat.setAttribute("PosNo", pt["PosNo"])
            feat.setAttribute("Event", pt["Event"])
            feat.setAttribute("DistCumulative", pt["DistCumulative"])
            feat.setAttribute("CableDistCumulative", pt["CableDistCumulative"])
            feat.setAttribute("ApproxDepth", pt["ApproxDepth"])
            feat.setAttribute("Remarks", pt["Remarks"])
            feat.setAttribute("ChartNo", pt.get("ChartNo", 0))
            feat.setAttribute("Latitude", pt["Lat"])
            feat.setAttribute("Longitude", pt["Lon"])
            feat.setAttribute("SourceFile", source_file_name)
            point_sink.addFeature(feat, QgsFeatureSink.FastInsert)

        # Create line features by pairing consecutive points with corresponding line data
        total_lines = min(len(point_list) - 1, len(line_list))
        for i in range(total_lines):
            pt1 = point_list[i]
            pt2 = point_list[i+1]
            line_attrs = line_list[i]
            geom_line = QgsGeometry.fromPolylineXY([QgsPointXY(pt1["Lon"], pt1["Lat"]),
                                                      QgsPointXY(pt2["Lon"], pt2["Lat"])])
            feat_line = QgsFeature(line_fields)
            feat_line.setGeometry(geom_line)
            feat_line.setAttribute("FromPos", pt1["PosNo"])
            feat_line.setAttribute("ToPos", pt2["PosNo"])
            feat_line.setAttribute("Bearing", line_attrs.get("Bearing"))
            feat_line.setAttribute("DistBetweenPos", line_attrs.get("DistBetweenPos"))
            feat_line.setAttribute("Slack", line_attrs.get("Slack"))
            feat_line.setAttribute("CableDistBetweenPos", line_attrs.get("CableDistBetweenPos"))
            feat_line.setAttribute("CableCode", line_attrs.get("CableCode"))
            feat_line.setAttribute("FiberPair", line_attrs.get("FiberPair"))
            feat_line.setAttribute("CableType", line_attrs.get("CableType"))
            feat_line.setAttribute("LayDirection", line_attrs.get("LayDirection"))
            feat_line.setAttribute("LayVessel", line_attrs.get("LayVessel"))
            feat_line.setAttribute("ProtectionMethod", line_attrs.get("ProtectionMethod"))
            feat_line.setAttribute("DateInstalled", line_attrs.get("DateInstalled"))
            feat_line.setAttribute("TargetBurialDepth", line_attrs.get("TargetBurialDepth"))
            feat_line.setAttribute("BurialDepth", line_attrs.get("BurialDepth"))
            feat_line.setAttribute("TerritorialWater", line_attrs.get("TerritorialWater"))
            feat_line.setAttribute("EEZ", line_attrs.get("EEZ"))
            feat_line.setAttribute("SourceFile", source_file_name)
            line_sink.addFeature(feat_line, QgsFeatureSink.FastInsert)

        # Provide summary information
        feedback.pushInfo(f"RPL import completed successfully:")
        feedback.pushInfo(f"- Points processed: {len(point_list)}")
        feedback.pushInfo(f"- Lines processed: {len(line_list)}")
        feedback.pushInfo(f"- Last row processed: {start_row + processed - 1}")

        # Set up dynamic renaming using post-processor, keeping references via self
        self.point_renamer = Renamer(f"{filename}_Points")
        context.layerToLoadOnCompletionDetails(point_sink_id).setPostProcessor(self.point_renamer)

        self.line_renamer = Renamer(f"{filename}_Lines")
        context.layerToLoadOnCompletionDetails(line_sink_id).setPostProcessor(self.line_renamer)

        return {
            self.OUTPUT_POINTS: point_sink_id,
            self.OUTPUT_LINES: line_sink_id
        }

class Renamer(QgsProcessingLayerPostProcessorInterface):
    def __init__(self, layer_name):
        self.name = layer_name
        super().__init__()

    def postProcessLayer(self, layer, context, feedback):
        layer.setName(self.name)