# -*- coding: utf-8 -*-
"""Pure-Python checks for the RPL import core (no QGIS required).

Run directly with any Python 3: ``python tests/test_rpl_import_core.py``.
Uses the plugin's bundled openpyxl (lib/) to build small workbook fixtures in
a temp folder, then exercises reading, detection, parsing, source-row
integrity, units, and validation diagnostics.
"""

from __future__ import annotations

import importlib.util
import os
import sys
import tempfile
from pathlib import Path
from typing import List

ROOT = Path(__file__).resolve().parents[1]
LIB = ROOT / "lib"
if str(LIB) not in sys.path:
    sys.path.insert(0, str(LIB))

# Load the rpl_import package by path (plugin folder name has hyphens).
_spec = importlib.util.spec_from_file_location(
    "sct_rpl_import", ROOT / "rpl_import" / "__init__.py",
    submodule_search_locations=[str(ROOT / "rpl_import")],
)
_pkg = importlib.util.module_from_spec(_spec)
sys.modules[_spec.name] = _pkg
_spec.loader.exec_module(_pkg)

from sct_rpl_import import coords as C            # noqa: E402
from sct_rpl_import import detect as D            # noqa: E402
from sct_rpl_import import model as M             # noqa: E402
from sct_rpl_import import parser as P            # noqa: E402
from sct_rpl_import import reader as R            # noqa: E402
from sct_rpl_import import validate as V          # noqa: E402

from openpyxl import Workbook                     # noqa: E402
from openpyxl import LXML as OPENPYXL_USES_LXML   # noqa: E402


def _result(name: str, ok: bool, detail: str = "") -> bool:
    print("[%s] %s%s" % ("PASS" if ok else "FAIL", name,
                         (" — " + detail) if detail else ""))
    return bool(ok)


def _tmp(name: str) -> str:
    folder = tempfile.mkdtemp(prefix="rpl_import_test_")
    return os.path.join(folder, name)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------
def _canonical_alternating(path: str) -> None:
    """Industry-style alternating workbook: title, merged 2-row header with
    route/cable groups, split DDM, an extra column, and a decoy sheet."""
    wb = Workbook()
    decoy = wb.active
    decoy.title = "Notes"
    decoy["A1"] = "Project notes, no RPL here"

    ws = wb.create_sheet("RPL")
    ws["A1"] = "ACME Cable System — Route Position List"
    # header rows 3-4 (merged group titles above field names)
    headers_row3 = ["Pos", "Event", "Latitude", "", "", "Longitude", "", "",
                    "Course", "Route", "", "Slack", "Cable", "", "Depth",
                    "Cable Type", "Remarks", "Chart", "Zone"]
    headers_row4 = ["No", "", "Deg", "Min", "N/S", "Deg", "Min", "E/W",
                    "(deg)", "Dist (km)", "Total (km)", "(%)",
                    "Dist (km)", "Total (km)", "(m)", "", "", "No", ""]
    for col, text in enumerate(headers_row3, start=1):
        if text:
            ws.cell(row=3, column=col, value=text)
    for col, text in enumerate(headers_row4, start=1):
        if text:
            ws.cell(row=4, column=col, value=text)
    ws.merge_cells("C3:E3")
    ws.merge_cells("F3:H3")
    ws.merge_cells("J3:K3")
    ws.merge_cells("M3:N3")

    # data: points on rows 5,7,9,11 — segments on 6,8,10
    # ~0.01 deg lat ≈ 1.112 km per span, due north from 50N 1W
    points = [
        (1, "BMH A", 50, 0.0, "N", 1, 0.0, "W", 0.0, 0.0, 12.0, "A1", "UK"),
        (2, "AC1", 50, 0.6, "N", 1, 0.0, "W", 1.112, 1.134, 55.0, "A1", "UK"),
        (3, "JT1", 50, 1.2, "N", 1, 0.0, "W", 2.224, 2.268, 87.0, "A2", "UK"),
        (4, "EOC", 50, 1.8, "N", 1, 0.0, "W", 3.336, 3.402, 120.0, "A2", "UK"),
    ]
    segments = [
        (0.0, 1.112, 2.0, 1.134, "LW-A", "Seg 1 remark"),
        (0.0, 1.112, 2.0, 1.134, "LW-A", ""),
        (0.0, 1.112, 2.0, 1.134, "DA-B", "Seg 3 remark"),
    ]
    row = 5
    for i, pt in enumerate(points):
        (pos, event, latd, latm, lath, lond, lonm, lonh,
         kp, cable_cum, depth, chart, zone) = pt
        ws.cell(row=row, column=1, value=pos)
        ws.cell(row=row, column=2, value=event)
        ws.cell(row=row, column=3, value=latd)
        ws.cell(row=row, column=4, value=latm)
        ws.cell(row=row, column=5, value=lath)
        ws.cell(row=row, column=6, value=lond)
        ws.cell(row=row, column=7, value=lonm)
        ws.cell(row=row, column=8, value=lonh)
        ws.cell(row=row, column=11, value=kp)
        ws.cell(row=row, column=14, value=cable_cum)
        ws.cell(row=row, column=15, value=depth)
        ws.cell(row=row, column=18, value=chart)
        ws.cell(row=row, column=19, value=zone)
        if i < len(segments):
            course, dist, slack, cable, ctype, remark = segments[i]
            srow = row + 1
            ws.cell(row=srow, column=9, value=course)
            ws.cell(row=srow, column=10, value=dist)
            ws.cell(row=srow, column=12, value=slack)
            ws.cell(row=srow, column=13, value=cable)
            ws.cell(row=srow, column=16, value=ctype)
            ws.cell(row=srow, column=17, value=remark)
        row += 2
    ws.cell(row=row + 1, column=2, value="End of RPL — footer notes")
    wb.save(path)


def _flat_csv(path: str) -> None:
    lines = [
        "Pos No,Event,Latitude,Longitude,KP (km),Bearing,Dist Between (km),Slack (%),Remarks",
        '1,BMH,"50° 0.000\' N","1° 0.000\' W",0.0,,,,start',
        '2,AC1,"50° 0.600\' N","1° 0.000\' W",1.112,0.0,1.112,2.0,',
        '3,EOC,"50° 1.200\' N","1° 0.000\' W",2.224,0.0,1.112,2.0,end',
    ]
    with open(path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(lines))


def _decimal_degrees_csv(path: str) -> None:
    lines = [
        "Pos,Event,Lat (dd),Lon (dd),KP (km)",
        "1,Start,50.0,-1.0,0.0",
        "2,Mid,50.01,-1.0,1.112",
        "3,End,50.02,-1.0,2.224",
    ]
    with open(path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(lines))


# ---------------------------------------------------------------------------
# Coordinate parser checks
# ---------------------------------------------------------------------------
def test_coord_parsers() -> bool:
    ok = OPENPYXL_USES_LXML is False
    value, reason = C.parse_split_ddm(50, 30.0, "N", C.AXIS_LAT)
    ok &= reason is None and abs(value - 50.5) < 1e-9
    value, reason = C.parse_split_ddm(1, 30.0, "W", C.AXIS_LON)
    ok &= reason is None and abs(value + 1.5) < 1e-9
    ok &= C.parse_split_ddm(50, 61.0, "N", C.AXIS_LAT)[1] == C.R_MINUTES_RANGE
    ok &= C.parse_split_ddm(50, 30.0, "E", C.AXIS_LAT)[1] == C.R_WRONG_AXIS_HEMISPHERE
    ok &= C.parse_split_ddm(95, 30.0, "N", C.AXIS_LAT)[1] == C.R_OUT_OF_RANGE
    ok &= C.parse_split_ddm(None, None, "", C.AXIS_LAT)[1] == C.R_EMPTY

    for text, expected in (
            ("50° 12.345' N", 50 + 12.345 / 60),
            ("50-12.345N", 50 + 12.345 / 60),
            ("N50 12.345", 50 + 12.345 / 60),
            ("50 12.345 N", 50 + 12.345 / 60)):
        value, reason = C.parse_ddm_text(text, C.AXIS_LAT)
        ok &= reason is None and abs(value - expected) < 1e-9
    value, reason = C.parse_ddm_text("1° 30.0' W", C.AXIS_LON)
    ok &= reason is None and abs(value + 1.5) < 1e-9
    value, reason = C.parse_ddm_text("50° 12' 30\" N", C.AXIS_LAT)  # strict DMS
    ok &= reason is None and abs(value - (50 + 12 / 60 + 30 / 3600)) < 1e-9
    ok &= C.parse_ddm_text("hello", C.AXIS_LAT)[1] == C.R_UNRECOGNISED
    ok &= C.parse_ddm_text("50° 75.0' N", C.AXIS_LAT)[1] == C.R_MINUTES_RANGE

    value, reason = C.parse_decimal_degrees(-50.25, C.AXIS_LAT)
    ok &= reason is None and value == -50.25
    value, reason = C.parse_decimal_degrees("50.25 S", C.AXIS_LAT)
    ok &= reason is None and value == -50.25
    ok &= C.parse_decimal_degrees("-50.25 N", C.AXIS_LAT)[1] == C.R_UNRECOGNISED
    ok &= C.parse_decimal_degrees(200.0, C.AXIS_LON)[1] == C.R_OUT_OF_RANGE
    return _result("coordinate parsers (split DDM / DDM text / decimal)", ok)


# ---------------------------------------------------------------------------
# Canonical alternating workbook
# ---------------------------------------------------------------------------
def _detect_canonical():
    path = _tmp("canonical.xlsx")
    _canonical_alternating(path)
    grids = R.load_sample_grids(path)
    results = D.score_sheets(grids)
    return path, results


def test_sheet_scoring_and_detection() -> bool:
    path, results = _detect_canonical()
    best = results[0]
    ok = best.profile.sheet == "RPL"
    ok &= best.position_count == 4
    profile = best.profile
    ok &= profile.coord_encoding == M.COORD_SPLIT_DDM
    ok &= profile.layout == M.LAYOUT_ALTERNATING
    ok &= profile.data_start_row == 5 and profile.data_end_row == 11
    mapping = profile.mapping
    ok &= mapping.get(M.PF_LAT_DEG) == 3 and mapping.get(M.PF_LAT_MIN) == 4
    ok &= mapping.get(M.PF_LAT_HEMI) == 5 and mapping.get(M.PF_LON_HEMI) == 8
    ok &= mapping.get(M.PF_POS_NO) == 1 and mapping.get(M.PF_EVENT) == 2
    ok &= mapping.get(M.SF_BEARING) == 9      # decoy "Course" maps to bearing
    ok &= mapping.get(M.SF_DIST) == 10
    ok &= mapping.get(M.PF_DIST_CUM) == 11
    ok &= mapping.get(M.SF_SLACK) == 12
    ok &= mapping.get(M.SF_CABLE_DIST) == 13
    ok &= mapping.get(M.PF_CABLE_DIST_CUM) == 14
    ok &= mapping.get(M.PF_DEPTH) == 15
    ok &= mapping.get(M.SF_CABLE_TYPE) == 16
    ok &= mapping.get(M.PF_REMARKS) == 17
    ok &= mapping.get(M.PF_CHART_NO) == 18
    ok &= best.confidence > 0.5
    detail = "" if ok else f"mapping={mapping}, rows={profile.data_start_row}-{profile.data_end_row}, enc={profile.coord_encoding}"
    return _result("worksheet scoring + detection on canonical workbook", ok, detail)


def test_parse_canonical() -> bool:
    path, results = _detect_canonical()
    profile = results[0].profile
    grid = R.load_grid(path, profile.sheet)
    doc, diags = P.parse(grid, profile)
    ok = len(doc.points) == 4 and len(doc.segments) == 3
    ok &= not M.has_errors(diags)
    p0, p1 = doc.points[0], doc.points[1]
    ok &= p0.pos_no == 1 and p0.event == "BMH A"
    ok &= abs(p0.lat - 50.0) < 1e-9 and abs(p0.lon + 1.0) < 1e-9
    ok &= abs(p1.lat - 50.01) < 1e-9
    ok &= p0.dist_cum_km == 0.0 and abs(p1.cable_dist_cum_km - 1.134) < 1e-9
    ok &= p0.depth_m == 12.0 and p0.chart_no == "A1"
    seg0 = doc.segments[0]
    ok &= abs(seg0.dist_km - 1.112) < 1e-9 and seg0.slack_pct == 2.0
    ok &= seg0.cable_type == "LW-A"
    ok &= seg0.source_row == 6 and doc.points[3].source_row == 11
    # unidentified columns start excluded and never leak into the model
    ok &= "Zone" not in p0.extras
    # validation on the clean doc: no errors, no distance mismatches
    findings = V.validate(doc)
    ok &= not M.has_errors(findings)
    ok &= not [f for f in findings if f.rule_id == "rpl_import.distance_mismatch"]
    return _result("parse canonical workbook (values, extras, source rows)", ok,
                   "" if ok else f"diags={[d.rule_id for d in diags]}")


def test_invalid_middle_point_blocks_and_never_shifts() -> bool:
    path = _tmp("badmid.xlsx")
    _canonical_alternating(path)
    # Detect/confirm the profile on the clean workbook (the wizard flow),
    # then corrupt one middle point's latitude and re-parse.
    profile = D.score_sheets(R.load_sample_grids(path))[0].profile
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["RPL"]
    ws.cell(row=7, column=4, value="garbage")   # break point 2's latitude
    wb.save(path)
    grid = R.load_grid(path, "RPL")
    doc, diags = P.parse(grid, profile)
    ok = M.has_errors(diags)
    ok &= any(d.rule_id == "rpl_import.point.invalid_coordinates" and d.row == 7
              for d in diags)
    # Slot structure intact: 4 point slots, 3 segments, attributes unshifted.
    ok &= len(doc.points) == 4 and len(doc.segments) == 3
    ok &= doc.segments[2].cable_type == "DA-B"          # still on span 3->4
    ok &= doc.segments[2].source_row == 10
    ok &= doc.points[2].event == "JT1"
    return _result("invalid middle point blocks import; segments never shift", ok)


# ---------------------------------------------------------------------------
# Flat layouts
# ---------------------------------------------------------------------------
def test_flat_csv_ddm_text_arriving() -> bool:
    path = _tmp("flat.csv")
    _flat_csv(path)
    grids = R.load_sample_grids(path)
    best = D.score_sheets(grids)[0]
    profile = best.profile
    ok = profile.coord_encoding == M.COORD_DDM_TEXT
    ok &= profile.layout == M.LAYOUT_FLAT
    ok &= profile.data_start_row == 2 and profile.data_end_row == 4
    grid = R.load_grid(path)
    doc, diags = P.parse(grid, profile)
    ok &= len(doc.points) == 3 and len(doc.segments) == 2
    ok &= not M.has_errors(diags)
    ok &= abs(doc.points[0].lat - 50.0) < 1e-9
    ok &= abs(doc.points[2].lon + 1.0) < 1e-9
    # arriving semantics: row 3's span data belongs to segment 0 (points 0->1)
    ok &= doc.segments[0].slack_pct == 2.0 and doc.segments[0].source_row == 3
    findings = V.validate(doc)
    ok &= not M.has_errors(findings)
    return _result("flat CSV with DDM text, arriving spans", ok,
                   "" if ok else f"enc={profile.coord_encoding} layout={profile.layout} rows={profile.data_start_row}-{profile.data_end_row}")


def test_flat_departing_semantics() -> bool:
    path = _tmp("flatdep.csv")
    _flat_csv(path)
    profile = D.score_sheets(R.load_sample_grids(path))[0].profile
    profile.flat_semantics = M.FLAT_DEPARTING
    grid = R.load_grid(path)
    doc, diags = P.parse(grid, profile)
    ok = len(doc.points) == 3 and len(doc.segments) == 2
    # departing: row 2's (empty) span joins 0->1; row 3's data joins 1->2
    ok &= doc.segments[1].slack_pct == 2.0 and doc.segments[1].source_row == 3
    ok &= doc.segments[0].slack_pct is None
    # last row's departing values (row 4 has slack too) are reported ignored
    ok &= any(d.rule_id == "rpl_import.layout.trailing_segment" for d in diags)
    return _result("flat departing-span semantics", ok)


def test_decimal_degrees_detection() -> bool:
    path = _tmp("dd.csv")
    _decimal_degrees_csv(path)
    best = D.score_sheets(R.load_sample_grids(path))[0]
    profile = best.profile
    ok = profile.coord_encoding == M.COORD_DECIMAL_DEGREES
    grid = R.load_grid(path)
    doc, diags = P.parse(grid, profile)
    ok &= len(doc.points) == 3 and not M.has_errors(diags)
    ok &= abs(doc.points[1].lon + 1.0) < 1e-9 and abs(doc.points[1].lat - 50.01) < 1e-9
    return _result("signed decimal degrees behind lat/lon headers", ok)


def test_grouped_headers_prefer_split_coordinates() -> bool:
    rows = [
        ["Pos", "Event", "Latitude", None, None, "Longitude", None, None,
         "Decimal Latitude (degrees)", "Radians Latitude", "Sin Latitude"],
        ["No.", None, None, None, None, None, None, None, None, None, None],
        [None] * 11,
        [None] * 11,
        [None] * 11,
        [None] * 11,
        [1, "BU SIN", -9, 36.1115, "S", 101, 30.1866, "E",
         -9.60185833333332, -0.167583, -0.1668],
        [None] * 11,
        [2, "AC", -9, 39.075, "S", 101, 29.2593, "E",
         -9.65125, -0.168445, -0.16765],
        [None] * 11,
        [3, "AC", -9, 42.0, "S", 101, 28.0, "E",
         -9.7, -0.169297, -0.16849],
    ]
    grid = R.SourceGrid(sheet="RPL", rows=rows, n_cols=11)
    result = D.detect(grid)
    profile = result.profile
    ok = profile.header_rows == [1, 2]
    ok &= profile.coord_encoding == M.COORD_SPLIT_DDM
    ok &= profile.mapping.get(M.PF_LAT_DEG) == 3
    ok &= profile.mapping.get(M.PF_LAT_MIN) == 4
    ok &= profile.mapping.get(M.PF_LAT_HEMI) == 5
    ok &= profile.mapping.get(M.PF_LON_DEG) == 6
    ok &= profile.mapping.get(M.PF_LON_MIN) == 7
    ok &= profile.mapping.get(M.PF_LON_HEMI) == 8
    ok &= "latitude" in result.header_texts[3]
    ok &= "longitude" in result.header_texts[6]
    ok &= "no" not in result.header_texts[3]
    ok &= "no" not in result.header_texts[6]
    return _result("grouped headers + values prefer split DDM coordinates", ok,
                   "" if ok else f"encoding={profile.coord_encoding} mapping={profile.mapping} headers={result.header_texts[:8]}")


def _industry_ddm_rows(with_hemis: bool = True) -> list:
    """Sheet shaped like real industry RPLs: banner + grouped 'Latitude' /
    'Longitude' headers over unlabelled deg/min/hemi sub-columns, derived
    decimal/radians/sin columns alongside, and a footer whose stray numbers
    land in the coordinate columns' sheet-wide stats."""
    n = 14

    def pad(cells):
        return cells + [None] * (n - len(cells))

    rows = [
        pad(["Straits RPL", None, None, None, None, None, "Issue", 3]),
        pad(["Pos", "Event", "Latitude", None, None, "Longitude", None, None,
             "Decimal", "Radians", "Sin", "Bearing", "Dist. Between Pos.",
             "Decimal"]),
        pad(["No.", None, None, None, None, None, None, None,
             "Latitude", "Latitude", "Latitude", None, "km", "Longitude"]),
        pad([None, None, None, None, None, None, None, None,
             "(degrees)", None, None, None, None, "(degrees)"]),
        pad([None] * n),
        pad([None] * n),
    ]
    import math
    lat = [(-9, 36.1115), (-9, 39.075), (-9, 42.0), (-10, 10.3268), (-10, 31.2743)]
    lon = [(101, 30.1866), (101, 29.2593), (102, 9.2495), (102, 32.1752), (103, 2.3096)]
    for i in range(5):
        la_d, la_m = lat[i]
        lo_d, lo_m = lon[i]
        dd_lat = -(abs(la_d) + la_m / 60.0)
        dd_lon = lo_d + lo_m / 60.0
        hemi_lat = "S" if with_hemis else None
        hemi_lon = "E" if with_hemis else None
        rows.append(pad([i + 1, "AC", la_d, la_m, hemi_lat, lo_d, lo_m, hemi_lon,
                         dd_lat, math.radians(dd_lat), math.sin(math.radians(dd_lat)),
                         None, None, dd_lon]))
        rows.append(pad([None, None, None, None, None, None, None, None,
                         None, None, None, 135.2, 12.345, None]))
    rows.append(pad([None] * n))
    rows.append(pad(["Notes:", None, 1, "Datum WGS 84 (2024)"]))
    rows.append(pad([None, None, 2, "Total route length", None, 1234.5, "km"]))
    return rows


def test_split_ddm_survives_footer_contamination() -> bool:
    grid = R.SourceGrid(sheet="RPL", rows=_industry_ddm_rows(), n_cols=14)
    result = D.detect(grid)
    profile = result.profile
    ok = profile.coord_encoding == M.COORD_SPLIT_DDM
    ok &= profile.mapping.get(M.PF_LAT_DEG) == 3
    ok &= profile.mapping.get(M.PF_LAT_MIN) == 4
    ok &= profile.mapping.get(M.PF_LAT_HEMI) == 5
    ok &= profile.mapping.get(M.PF_LON_DEG) == 6
    ok &= profile.mapping.get(M.PF_LON_MIN) == 7
    ok &= profile.mapping.get(M.PF_LON_HEMI) == 8
    doc, diags = P.parse(grid, profile)
    ok &= len(doc.points) == 5 and not M.has_errors(diags)
    if doc.points:
        ok &= abs(doc.points[0].lat - -(9 + 36.1115 / 60.0)) < 1e-9
        ok &= abs(doc.points[0].lon - (101 + 30.1866 / 60.0)) < 1e-9
    return _result(
        "split DDM detected despite footer/banner numbers in scan", ok,
        "" if ok else f"encoding={profile.coord_encoding} mapping={profile.mapping}")


def test_decimal_fallback_prefers_true_decimal_columns() -> bool:
    # Same sheet but the hemisphere columns are empty, so split DDM cannot be
    # detected. The decimal fallback must then pick the actual decimal
    # degrees columns (fractional, 'decimal ... (degrees)' headers) — never
    # the integer degree/minute sub-columns under the grouped headers, and
    # never the radians/sin derived columns.
    grid = R.SourceGrid(sheet="RPL", rows=_industry_ddm_rows(with_hemis=False),
                        n_cols=14)
    profile = D.detect(grid).profile
    ok = profile.coord_encoding == M.COORD_DECIMAL_DEGREES
    ok &= profile.mapping.get(M.PF_LAT_TEXT) == 9
    ok &= profile.mapping.get(M.PF_LON_TEXT) == 14
    return _result(
        "decimal fallback prefers true decimal columns over deg/min/trig", ok,
        "" if ok else f"encoding={profile.coord_encoding} mapping={profile.mapping}")


def test_split_ddm_hemisphere_first_order() -> bool:
    rows = [
        ["Pos", "Lat", None, None, "Lon", None, None],
        [1, "N", 50, 10.5, "W", 1, 20.25],
        [2, "N", 50, 12.0, "W", 1, 22.5],
        [3, "N", 50, 13.5, "W", 1, 24.75],
    ]
    grid = R.SourceGrid(sheet="RPL", rows=rows, n_cols=7)
    profile = D.detect(grid).profile
    ok = profile.coord_encoding == M.COORD_SPLIT_DDM
    ok &= profile.mapping.get(M.PF_LAT_HEMI) == 2
    ok &= profile.mapping.get(M.PF_LAT_DEG) == 3
    ok &= profile.mapping.get(M.PF_LAT_MIN) == 4
    ok &= profile.mapping.get(M.PF_LON_HEMI) == 5
    ok &= profile.mapping.get(M.PF_LON_DEG) == 6
    ok &= profile.mapping.get(M.PF_LON_MIN) == 7
    return _result(
        "split DDM with hemisphere-first column order", ok,
        "" if ok else f"encoding={profile.coord_encoding} mapping={profile.mapping}")


def test_merged_header_industry_layout() -> bool:
    """Layout from real industry RPLs: 'Latitude'/'Longitude' merged over
    deg/min/hemi triples (the reader replicates merge anchors into every
    covered cell), grouped 'Distance (km)' / 'Cable Distance (km)' headers
    with Between/Cumulative sub-headers, a run of hidden empty columns, and
    trailing engineering columns."""
    n = 31

    def pad(cells):
        return cells + [None] * (n - len(cells))

    gap = [None] * 12  # hidden columns I..T
    head = ["Pos No.", "Event", "Latitude", "Latitude", "Latitude",
            "Longitude", "Longitude", "Longitude"] + gap + ["Bearing °T"]
    rows = [
        pad(head + ["Distance (km)", "Distance (km)", "Slack %",
                    "Cable Distance (km)", "Cable Distance (km)", "Cable Code",
                    "Fiber Pair", "Cable Type", "Approx Depth (m)", "Remarks"]),
        pad(head + ["Between Positions", "Cumulative Total", "Slack %",
                    "Between Positions", "Cumulative Total", "Cable Code",
                    "Fiber Pair", "Cable Type", "Approx Depth (m)", "Remarks"]),
    ]
    lat = [(9, 36.1115), (9, 39.075), (9, 42.0), (10, 10.3268)]
    lon = [(101, 30.1866), (101, 29.2593), (102, 9.2495), (102, 32.1752)]
    total = 0.0
    for i in range(4):
        la_d, la_m = lat[i]
        lo_d, lo_m = lon[i]
        rows.append(pad([i + 1, "AC", la_d, la_m, "S", lo_d, lo_m, "E"]
                        + gap + [None, None, None, None, None, None,
                                 None, None, None, 1520.0 + i, "note"]))
        if i < 3:
            seg = 12.345
            total += seg
            rows.append(pad([None] * 8 + gap
                            + [135.2, seg, round(total, 3), 1.5,
                               round(seg * 1.015, 3), round(total * 1.015, 3),
                               "A1", "1-4", "LW" if i else "DA",
                               None, None]))
    grid = R.SourceGrid(sheet="RPL", rows=rows, n_cols=n)
    result = D.detect(grid)
    profile = result.profile
    ok = profile.coord_encoding == M.COORD_SPLIT_DDM
    ok &= profile.mapping.get(M.PF_LAT_DEG) == 3
    ok &= profile.mapping.get(M.PF_LAT_MIN) == 4
    ok &= profile.mapping.get(M.PF_LAT_HEMI) == 5
    ok &= profile.mapping.get(M.PF_LON_DEG) == 6
    ok &= profile.mapping.get(M.PF_LON_MIN) == 7
    ok &= profile.mapping.get(M.PF_LON_HEMI) == 8
    ok &= profile.mapping.get(M.SF_BEARING) == 21
    ok &= profile.mapping.get(M.SF_DIST) == 22
    ok &= profile.mapping.get(M.PF_DIST_CUM) == 23
    ok &= profile.mapping.get(M.SF_SLACK) == 24
    ok &= profile.mapping.get(M.SF_CABLE_DIST) == 25
    ok &= profile.mapping.get(M.PF_CABLE_DIST_CUM) == 26
    ok &= profile.mapping.get(M.SF_CABLE_CODE) == 27
    ok &= profile.mapping.get(M.SF_FIBER_PAIR) == 28
    ok &= profile.mapping.get(M.SF_CABLE_TYPE) == 29
    ok &= profile.mapping.get(M.PF_DEPTH) == 30
    ok &= profile.mapping.get(M.PF_REMARKS) == 31
    ok &= profile.distance_unit == "km"
    doc, diags = P.parse(grid, profile)
    ok &= len(doc.points) == 4 and not M.has_errors(diags)
    return _result(
        "merged-header industry layout fully mapped", ok,
        "" if ok else f"encoding={profile.coord_encoding} mapping={profile.mapping}")


def test_detect_coordinate_columns_for_encoding() -> bool:
    # Split triples exist: constrained split detection finds them.
    grid = R.SourceGrid(sheet="RPL", rows=_industry_ddm_rows(), n_cols=14)
    split = D.detect_coordinate_columns(grid, M.COORD_SPLIT_DDM)
    ok = split == {M.PF_LAT_DEG: 3, M.PF_LAT_MIN: 4, M.PF_LAT_HEMI: 5,
                   M.PF_LON_DEG: 6, M.PF_LON_MIN: 7, M.PF_LON_HEMI: 8}
    # Same sheet: constrained decimal detection targets the decimal columns.
    decimal = D.detect_coordinate_columns(grid, M.COORD_DECIMAL_DEGREES)
    ok &= decimal == {M.PF_LAT_TEXT: 9, M.PF_LON_TEXT: 14}
    # No hemisphere letters: split constrained detection reports nothing.
    bare = R.SourceGrid(sheet="RPL", rows=_industry_ddm_rows(with_hemis=False),
                        n_cols=14)
    ok &= D.detect_coordinate_columns(bare, M.COORD_SPLIT_DDM) is None
    # Projected has no content signature.
    ok &= D.detect_coordinate_columns(grid, M.COORD_PROJECTED) is None
    return _result("encoding-constrained coordinate detection", ok)


def test_split_cells_with_unit_symbols() -> bool:
    # Text cells carrying degree/minute symbols still parse and detect.
    ok = abs(C.parse_split_ddm("9°", "36.1115'", "S", C.AXIS_LAT)[0]
             - -(9 + 36.1115 / 60.0)) < 1e-9
    rows = [
        ["Pos", "Latitude", None, None, "Longitude", None, None],
        [1, "9°", "36.1115'", "S", "101°", "30.1866'", "E"],
        [2, "9°", "39.0750'", "S", "101°", "29.2593'", "E"],
        [3, "9°", "42.0000'", "S", "102°", "9.2495'", "E"],
    ]
    grid = R.SourceGrid(sheet="RPL", rows=rows, n_cols=7)
    profile = D.detect(grid).profile
    ok &= profile.coord_encoding == M.COORD_SPLIT_DDM
    ok &= profile.mapping.get(M.PF_LAT_DEG) == 2
    ok &= profile.mapping.get(M.PF_LAT_MIN) == 3
    ok &= profile.mapping.get(M.PF_LAT_HEMI) == 4
    doc, diags = P.parse(grid, profile)
    ok &= len(doc.points) == 3 and not M.has_errors(diags)
    return _result("split cells with °/' unit symbols", ok,
                   "" if ok else f"encoding={profile.coord_encoding} mapping={profile.mapping}")


def test_canonical_fields_beat_derived_navigation_columns() -> bool:
    rows = [
        ["Pos No.", "Event", "Latitude", None, None, "Longitude", None, None,
         "Course", "Distance in miles (6087 ft)", "Bearing °T",
         "Distance (km) Between Positions", "Cumulative Total", "Slack %"],
        [1, "Start", 50, 0.0, "N", 1, 0.0, "W",
         None, None, None, None, 0.0, None],
        [None, None, None, None, None, None, None, None,
         0.01, 0.60, 0.0, 1.112, None, 2.0],
        [2, "Mid", 50, 0.6, "N", 1, 0.0, "W",
         None, None, None, None, 1.112, None],
        [None, None, None, None, None, None, None, None,
         0.01, 0.60, 0.0, 1.112, None, 2.0],
        [3, "End", 50, 1.2, "N", 1, 0.0, "W",
         None, None, None, None, 2.224, None],
    ]
    grid = R.SourceGrid(sheet="RPL", rows=rows, n_cols=14)
    profile = D.detect(grid).profile
    ok = profile.mapping.get(M.SF_BEARING) == 11
    ok &= profile.mapping.get(M.SF_DIST) == 12
    ok &= profile.mapping.get(M.PF_DIST_CUM) == 13
    ok &= profile.mapping.get(M.SF_SLACK) == 14
    return _result("canonical fields beat derived navigation columns", ok,
                   "" if ok else f"mapping={profile.mapping}")


def test_undetected_sheet_starts_fully_excluded() -> bool:
    grid = R.SourceGrid(
        sheet="Notes", rows=[["Topic", "Owner"], ["Permits", "Team"]],
        n_cols=2)
    profile = D.detect(grid).profile
    ok = profile.mapping == {} and profile.excluded_columns == [1, 2]
    return _result("undetected sheet columns start excluded", ok,
                   "" if ok else f"excluded={profile.excluded_columns}")


# ---------------------------------------------------------------------------
# Units, profile round-trip, misc invariants
# ---------------------------------------------------------------------------
def test_unit_conversion_and_slack_ratio() -> bool:
    profile = M.ImportProfile()
    ok = profile.source_crs == "EPSG:4326"
    ok &= M.to_km(1500.0, "m") == 1.5
    ok &= abs(M.to_km(1.0, "nm") - 1.852) < 1e-12
    ok &= M.to_m(10.0, "ft") == 3.048
    ok &= M.slack_to_percent(1.02, True) is not None
    ok &= abs(M.slack_to_percent(1.02, True) - 2.0) < 1e-9
    ok &= M.slack_to_percent(2.0, False) == 2.0
    ok &= M.to_km(None, "m") is None and M.slack_to_percent(None, True) is None
    return _result("unit conversion helpers", ok)


def test_units_detected_from_headers() -> bool:
    path = _tmp("units.csv")
    lines = [
        "Pos,Event,Lat (dd),Lon (dd),KP (m),Dist Between (m)",
        "1,Start,50.0,-1.0,0,",
        "2,End,50.01,-1.0,1112,1112",
    ]
    with open(path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(lines))
    best = D.score_sheets(R.load_sample_grids(path))[0]
    ok = best.profile.distance_unit == "m"
    grid = R.load_grid(path)
    doc, _ = P.parse(grid, best.profile)
    ok &= doc.points[1].dist_cum_km == 1.112
    ok &= doc.segments[0].dist_km == 1.112
    findings = V.validate(doc)
    ok &= not [f for f in findings if f.rule_id == "rpl_import.distance_mismatch"]
    return _result("distance units read from headers and converted", ok)


def test_profile_json_round_trip_and_signature() -> bool:
    profile = M.ImportProfile(
        sheet="RPL", data_start_row=5, data_end_row=11,
        layout=M.LAYOUT_ALTERNATING, coord_encoding=M.COORD_SPLIT_DDM,
        mapping={M.PF_LAT_DEG: 3, M.PF_POS_NO: 1}, excluded_columns=[19],
        distance_unit="km", header_signature="abc")
    restored = M.ImportProfile.from_json(profile.to_json())
    ok = restored.mapping == profile.mapping
    ok &= restored.excluded_columns == [19]
    ok &= restored.data_end_row == 11
    sig1 = M.header_signature(["Pos No", "Event", "Latitude Deg"])
    sig2 = M.header_signature(["pos  no", "EVENT", "latitude-deg"])
    sig3 = M.header_signature(["Pos No", "Event", "Longitude Deg"])
    ok &= sig1 == sig2 and sig1 != sig3
    missing = M.ImportProfile(coord_encoding=M.COORD_SPLIT_DDM).required_missing()
    ok &= M.PF_LAT_DEG in missing and M.PF_LON_HEMI in missing
    dupes = M.ImportProfile(mapping={M.PF_POS_NO: 2, M.PF_EVENT: 2}).duplicate_assignments()
    ok &= dupes == {2: [M.PF_POS_NO, M.PF_EVENT]}
    return _result("profile JSON round-trip, signatures, mapping guards", ok)


def test_extra_field_names_collision_safe() -> bool:
    taken = set()
    a = M.extra_field_name("Zone", 19, taken)
    b = M.extra_field_name("Zone", 20, taken)
    c = M.extra_field_name("", 21, taken)
    d = M.extra_field_name("Slack", 22, taken)     # reserved canonical name
    ok = a == "Zone" and b == "Zone_2" and c == "col_21"
    ok &= d.lower() not in M.RESERVED_FIELD_NAMES
    return _result("extra column names deterministic and collision-safe", ok,
                   f"{a},{b},{c},{d}")


def test_non_integer_pos_no_and_chart_text() -> bool:
    path = _tmp("posno.csv")
    lines = [
        "Pos,Event,Lat (dd),Lon (dd),Chart",
        "1,Start,50.0,-1.0,GB1234",
        "2A,Mid,50.01,-1.0,GB1234",
        "3,End,50.02,-1.0,555",
    ]
    with open(path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(lines))
    best = D.score_sheets(R.load_sample_grids(path))[0]
    grid = R.load_grid(path)
    doc, diags = P.parse(grid, best.profile)
    ok = doc.points[0].chart_no == "GB1234"     # alphanumeric kept as text
    ok &= doc.points[1].pos_no is None and doc.points[1].pos_no_raw == "2A"
    ok &= any(d.rule_id == "rpl_import.point.non_integer_pos_no" for d in diags)
    ok &= not M.has_errors(diags)               # warning, not a blocker
    return _result("non-integer PosNo kept as evidence; ChartNo stays text", ok)


def test_notes_block_and_footer_are_outside_range() -> bool:
    path = _tmp("banner.csv")
    lines = [
        "Pos,Event,Lat (dd),Lon (dd)",
        "1,Start,50.0,-1.0",
        "2,Mid,50.01,-1.0",
        "--- section B ---,,,",
        "3,Mid2,50.02,-1.0",
        "4,End,50.03,-1.0",
        "Prepared by: someone,,,",
        "Checked by: someone else,,,",
    ]
    with open(path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(lines))
    best = D.score_sheets(R.load_sample_grids(path))[0]
    profile = best.profile
    ok = profile.data_start_row == 2 and profile.data_end_row == 6
    grid = R.load_grid(path)
    doc, diags = P.parse(grid, profile)
    ok &= len(doc.points) == 5          # banner row parsed as invalid point slot
    ok &= M.has_errors(diags)           # ...and blocks until resolved
    # user narrows the end row: banner row diagnostic remains, footer ignored
    return _result("notes banner inside data is surfaced; footer excluded", ok,
                   "" if ok else f"rows={profile.data_start_row}-{profile.data_end_row} pts={len(doc.points)}")


def test_full_sheet_range_detection_excludes_footer() -> bool:
    rows = [["Pos", "Event", "Lat (dd)", "Lon (dd)"]]
    for index in range(4500):
        rows.append([index + 1, "Route point", 50.0 + index / 100000.0, -1.0])
    rows.extend([
        ["Prepared by: someone", None, None, None],
        ["Checked by: someone else", None, None, None],
    ])
    grid = R.SourceGrid(sheet="RPL", rows=rows, n_cols=4)
    profile = D.detect(grid).profile
    ok = profile.data_start_row == 2 and profile.data_end_row == 4501
    return _result("full RPL row range retained; trailing notes excluded", ok,
                   "" if ok else f"rows={profile.data_start_row}-{profile.data_end_row}")


def test_manual_range_override_is_authoritative() -> bool:
    path = _tmp("override.csv")
    _flat_csv(path)
    profile = D.score_sheets(R.load_sample_grids(path))[0].profile
    profile.data_end_row = 3            # user trims the last row
    grid = R.load_grid(path)
    doc, diags = P.parse(grid, profile)
    ok = len(doc.points) == 2 and len(doc.segments) == 1
    ok &= doc.points[-1].source_row == 3
    return _result("manual data-range override is authoritative", ok)


def test_excluded_vs_preserved_extras() -> bool:
    path, results = _detect_canonical()
    profile = results[0].profile
    zone_col = 19
    ok = zone_col in profile.excluded_columns
    grid = R.load_grid(path, profile.sheet)
    doc, _ = P.parse(grid, profile)
    ok &= "Zone" not in doc.points[0].extras
    profile.excluded_columns.remove(zone_col)
    doc2, _ = P.parse(grid, profile)
    ok &= doc2.points[0].extras.get("Zone") == "UK"
    return _result("unidentified columns excluded until explicitly included", ok)


# ---------------------------------------------------------------------------
# Validation diagnostics
# ---------------------------------------------------------------------------
def _simple_doc():
    doc = M.ImportedRpl(sheet="T")
    for i, (lat, lon, kp) in enumerate(
            ((50.0, -1.0, 0.0), (50.01, -1.0, 1.112), (50.02, -1.0, 2.224))):
        doc.points.append(M.ImportPoint(
            seq=i, source_row=i + 2, pos_no=i + 1, lat=lat, lon=lon,
            dist_cum_km=kp))
    for i in range(2):
        doc.segments.append(M.ImportSegment(
            seq=i, source_row=i + 2, dist_km=1.112, slack_pct=2.0,
            cable_dist_km=1.134))
    return doc


def test_validation_rules() -> bool:
    ok = True

    doc = _simple_doc()
    ok &= not M.has_errors(V.validate(doc))

    bad = _simple_doc()
    bad.segments[0].dist_km = 5.0
    ok &= any(f.rule_id == "rpl_import.distance_mismatch"
              for f in V.validate(bad))

    kp = _simple_doc()
    kp.points[2].dist_cum_km = 0.5
    ok &= any(f.rule_id == "rpl_import.kp.non_monotonic" for f in V.validate(kp))

    dup = _simple_doc()
    dup.points[1].pos_no = 1
    ok &= any(f.rule_id == "rpl_import.pos_no.duplicate" for f in V.validate(dup))

    co = _simple_doc()
    co.points[1].lat, co.points[1].lon = 50.0, -1.0
    ok &= any(f.rule_id == "rpl_import.coincident_points" for f in V.validate(co))

    slack = _simple_doc()
    slack.segments[0].cable_dist_km = 9.9
    ok &= any(f.rule_id == "rpl_import.slack_inconsistent"
              for f in V.validate(slack))

    sign = _simple_doc()
    sign.points[1].lon = 1.0            # isolated E/W flip
    findings = V.validate(sign)
    ok &= any(f.rule_id == "rpl_import.coordinate_sign_outlier" for f in findings)

    # legitimate meridian crossing: near-zero magnitudes must NOT flag
    cross = _simple_doc()
    cross.points[0].lon = -0.02
    cross.points[1].lon = 0.01
    cross.points[2].lon = 0.05
    ok &= not any(f.rule_id == "rpl_import.coordinate_sign_outlier"
                  for f in V.validate(cross))

    cum = _simple_doc()
    cum.points[2].dist_cum_km = 9.0
    ok &= any(f.rule_id == "rpl_import.kp.cumulative_mismatch"
              for f in V.validate(cum))

    few = M.ImportedRpl(points=[M.ImportPoint(seq=0, lat=50.0, lon=-1.0)])
    ok &= M.has_errors(V.validate(few))
    return _result("validation rule battery", ok)


def test_uncached_formula_reported() -> bool:
    path = _tmp("formula.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "RPL"
    ws.append(["Pos", "Event", "Lat (dd)", "Lon (dd)", "KP (km)"])
    ws.append([1, "Start", 50.0, -1.0, 0.0])
    ws.append([2, "End", 50.01, -1.0, "=B99+1"])   # never calculated
    wb.save(path)
    best = D.score_sheets(R.load_sample_grids(path))[0]
    grid = R.load_grid(path, "RPL")
    doc, diags = P.parse(grid, best.profile)
    ok = any(d.rule_id == "rpl_import.cell.uncached_formula" and d.row == 3
             for d in diags)
    ok &= doc.points[1].dist_cum_km is None        # empty, never zero
    return _result("uncached formula cells surfaced, values left empty", ok)


def test_uncached_formula_ignored_in_excluded_column() -> bool:
    path = _tmp("formula_excluded.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "RPL"
    ws.append(["Pos", "Event", "Lat (dd)", "Lon (dd)", "Derived", "KP (km)"])
    ws.append([1, "Start", 50.0, -1.0, "=B99+1", 0.0])
    ws.append([2, "End", 50.01, -1.0, "=B99+2", "=B99+3"])
    wb.save(path)
    profile = D.score_sheets(R.load_sample_grids(path))[0].profile
    grid = R.load_grid(path, "RPL")
    _doc, diags = P.parse(grid, profile)
    formula_diags = [
        diag for diag in diags
        if diag.rule_id == "rpl_import.cell.uncached_formula"
    ]
    ok = 5 in profile.excluded_columns
    ok &= [(diag.row, diag.column) for diag in formula_diags] == [(3, 6)]
    return _result("excluded uncached formulas suppressed; mapped warning kept", ok,
                   "" if ok else f"excluded={profile.excluded_columns} warnings={[(d.row, d.column) for d in formula_diags]}")


def test_unsafe_xml_backend_refused() -> bool:
    import openpyxl

    original = openpyxl.LXML
    try:
        openpyxl.LXML = True
        try:
            R._require_openpyxl()
            ok = False
        except R.ReaderError as exc:
            ok = "Restart QGIS" in str(exc)
    finally:
        openpyxl.LXML = original
    return _result("unsafe preloaded XML backend refused safely", ok)


def run_all() -> List[bool]:
    results = [
        test_coord_parsers(),
        test_sheet_scoring_and_detection(),
        test_parse_canonical(),
        test_invalid_middle_point_blocks_and_never_shifts(),
        test_flat_csv_ddm_text_arriving(),
        test_flat_departing_semantics(),
        test_decimal_degrees_detection(),
        test_grouped_headers_prefer_split_coordinates(),
        test_split_ddm_survives_footer_contamination(),
        test_decimal_fallback_prefers_true_decimal_columns(),
        test_split_ddm_hemisphere_first_order(),
        test_merged_header_industry_layout(),
        test_detect_coordinate_columns_for_encoding(),
        test_split_cells_with_unit_symbols(),
        test_canonical_fields_beat_derived_navigation_columns(),
        test_undetected_sheet_starts_fully_excluded(),
        test_unit_conversion_and_slack_ratio(),
        test_units_detected_from_headers(),
        test_profile_json_round_trip_and_signature(),
        test_extra_field_names_collision_safe(),
        test_non_integer_pos_no_and_chart_text(),
        test_notes_block_and_footer_are_outside_range(),
        test_full_sheet_range_detection_excludes_footer(),
        test_manual_range_override_is_authoritative(),
        test_excluded_vs_preserved_extras(),
        test_validation_rules(),
        test_uncached_formula_reported(),
        test_uncached_formula_ignored_in_excluded_column(),
        test_unsafe_xml_backend_refused(),
    ]
    print("")
    print(f"{sum(results)}/{len(results)} passed")
    return results


if __name__ == "__main__":
    raise SystemExit(0 if all(run_all()) else 1)
