# -*- coding: utf-8 -*-
"""Alternating-row RPL sheet: the classic published-RPL presentation.

Positions and legs interleave exactly the way alternating-layout workbooks
arrive through the import wizard: a position row, then the leg row joining it
to the next position, ending on a position row (n points, n-1 legs).
Cumulative quantities (KP, cable distance) sit on position rows; between-
position quantities (bearing, distance, slack) sit on leg rows.

Pure python — no Qt/QGIS imports — so the dock table view and the CSV/XLSX
exporters share one builder and it stays unit-testable.
"""

from __future__ import annotations

from typing import Dict, List, Sequence, Tuple

ROW_POINT = "point"
ROW_LEG = "leg"

POINT_ATTR_ORDER = ["Remarks", "ChartNo", "PosNoText", "SourceFile"]
LEG_ATTR_ORDER = [
    "CableType", "CableCode", "FiberPair", "LayDirection", "LayVessel",
    "ProtectionMethod", "DateInstalled", "TargetBurialDepth", "BurialDepth",
    "TerritorialWater", "EEZ", "SourceFile",
]
ATTR_LABELS = {
    "CableType": "Cable type", "CableCode": "Cable code", "FiberPair": "Fibre pair",
    "LayDirection": "Lay direction", "LayVessel": "Lay vessel",
    "ProtectionMethod": "Protection method", "DateInstalled": "Date installed",
    "TargetBurialDepth": "Target burial depth", "BurialDepth": "Burial depth",
    "TerritorialWater": "Territorial water", "SourceFile": "Source file",
    "ChartNo": "Chart no.", "PosNoText": "Position text",
}


def attribute_keys(attr_rows, preferred) -> List[str]:
    """Union of attribute keys over rows, preferred order first."""
    found = []
    for attrs in attr_rows:
        for key in attrs:
            if key not in found:
                found.append(key)
    ordered = [key for key in preferred if key in found]
    ordered.extend(sorted([key for key in found if key not in ordered],
                          key=lambda value: value.lower()))
    return ordered


def _text(value) -> str:
    return "" if value is None else str(value)


def _number(value, decimals: int) -> str:
    return "" if value is None else f"{float(value):.{decimals}f}"


def build_sheet(model) -> Tuple[List[str], List[List[str]], List[str]]:
    """Return (headers, rows, kinds) for the alternating presentation.

    ``kinds[i]`` is ROW_POINT or ROW_LEG so callers can style leg rows.
    """
    point_attrs = attribute_keys((p.attrs for p in model.points), POINT_ATTR_ORDER)
    leg_attrs = attribute_keys((s.attrs for s in model.segments), LEG_ATTR_ORDER)
    # avoid two identical column labels when both sides carry e.g. SourceFile
    shared = set(point_attrs) & set(leg_attrs)

    def point_attr_label(key):
        label = ATTR_LABELS.get(key, key)
        return f"{label} (position)" if key in shared else label

    def leg_attr_label(key):
        label = ATTR_LABELS.get(key, key)
        return f"{label} (leg)" if key in shared else label

    headers = (
        ["Pos", "Event", "Latitude", "Longitude", "Depth (m)",
         "Bearing (deg)", "Dist (km)", "KP (km)",
         "Slack (%)", "Cable (km)", "Cable cum. (km)"]
        + [point_attr_label(key) for key in point_attrs]
        + [leg_attr_label(key) for key in leg_attrs]
    )
    blank_point = [""] * len(point_attrs)
    blank_leg = [""] * len(leg_attrs)

    rows: List[List[str]] = []
    kinds: List[str] = []
    for i, point in enumerate(model.points):
        rows.append(
            [_text(point.pos_no), point.event or "",
             f"{point.lat:.6f}", f"{point.lon:.6f}",
             _number(point.depth_m, 1),
             "", "", _number(point.dist_cum_km, 3),
             "", "", _number(point.cable_dist_cum_km, 3)]
            + [_text(point.attrs.get(key)) for key in point_attrs]
            + blank_leg
        )
        kinds.append(ROW_POINT)
        if i < len(model.segments):
            seg = model.segments[i]
            rows.append(
                ["", "", "", "", "",
                 _number(seg.bearing_deg, 1), _number(seg.dist_km, 4), "",
                 _number(seg.slack_pct, 3), _number(seg.cable_dist_km, 4), ""]
                + blank_point
                + [_text(seg.attrs.get(key)) for key in leg_attrs]
            )
            kinds.append(ROW_LEG)
    return headers, rows, kinds


def write_csv(path: str, headers: Sequence[str], rows: Sequence[Sequence[str]]) -> None:
    import csv

    # utf-8-sig so Excel opens degrees/accents correctly
    with open(path, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(list(headers))
        writer.writerows(list(row) for row in rows)


def write_xlsx(path: str, headers: Sequence[str], rows: Sequence[Sequence[str]],
               kinds: Sequence[str], title: str = "RPL") -> None:
    """Write an .xlsx with leg rows tinted. Raises ImportError without openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    book = Workbook()
    sheet = book.active
    # Excel sheet titles: max 31 chars, no []:*?/\
    safe = "".join(c for c in title if c not in "[]:*?/\\").strip() or "RPL"
    sheet.title = safe[:31]
    sheet.append(list(headers))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    leg_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2",
                           fill_type="solid")
    for row_values, kind in zip(rows, kinds):
        sheet.append([_typed(value) for value in row_values])
        if kind == ROW_LEG:
            for cell in sheet[sheet.max_row]:
                cell.fill = leg_fill
    sheet.freeze_panes = "A2"
    book.save(path)


def _typed(value: str):
    """Numbers as numbers in Excel; everything else stays text."""
    text = "" if value is None else str(value)
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        pass
    try:
        return float(text)
    except ValueError:
        return text
